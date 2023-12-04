import os
import requests
from credentials import *
import openpyxl as xl

class MondayAPI:
    def __init__(self):
        self.base_url = "https://api.monday.com/v2"
        self.api_key = API_KEY
        self.headers = {
            'Authorization' : self.api_key
            }
    
    def get_board_id(self, url):
        uri = url.split('/')
        return uri[-1]

    def get_groups(self, board_id):
        groups = []
        query = "query { boards" + f"(ids: {board_id})" + "{ groups { id title }}}"
        data = {'query': query}

        response = requests.post(self.base_url, headers=self.headers, json=data)
        for group in response.json()['data']['boards'][0]['groups']:
            groups.append([group['id'], group['title']])
        
        return groups
    
    def create_item(self, board_id, group_id, item_name):
        query = "mutation { create_item " + f"(board_id: {board_id}, group_id: {group_id}, item_name: {item_name})" + "{ id } }"
        data = {'query': query}

        response = requests.post(self.base_url, headers=self.headers, json=data)

        return response.json()['data']['create_item']['id']
    
    def change_column_value(self, board_id, item_id, column_id, value):
        query = "mutation ($value: String!) { change_simple_column_value " + f"(item_id: {item_id}, board_id: {board_id}, column_id: {column_id}, value: $value )" + " { id } }"
        qvars = {"value": f"{value}"}
        data = {'query': query, 'variables': qvars}

        response = requests.post(self.base_url, headers=self.headers, json=data)

        '''print(response)
        print(response.json())'''
    
    def create_sop(self, parent_item_id, sop_string, item_name, folder_selected):
        sops = sop_string.split(',')
        final_str = ''
        for sop in sops:
            ts = self.search_in_sop(item_name, folder_selected, sop.replace('"', '').replace(' ', ''))
            final_str += ts + ', '
        query = "mutation { create_subitem " + f'(parent_item_id: {parent_item_id}, item_name: "{final_str}") ' + "{ board { id } } }"
        data = {'query': query}

        response = requests.post(self.base_url, headers=self.headers, json=data)

    
    def board_columns(self, board_id):
        columns = []
        query = "query { boards " + f"(ids: {board_id}) " + "{columns { id title } } }"
        data = {'query': query}

        response = requests.post(self.base_url, headers=self.headers, json=data)

        for column in response.json()['data']['boards'][0]['columns']:
            columns.append([column['id'], column['title']])
        
        return columns
    
    def add_to_monday(self, board_id, group_id, item_name, column_qt_id, column_un_id, status_id, quantity_unit, sops, folder_selected, create=True, item_id=None, original_name=None):
        quantity_unit = quantity_unit.split(' ')
        if create:
            item_id = self.create_item(board_id, group_id, item_name)
        self.change_column_value(board_id, item_id, column_qt_id, quantity_unit[0])
        self.change_column_value(board_id, item_id, column_un_id, quantity_unit[1])
        self.set_status(status_id, item_id, board_id)
        if original_name == None:
            self.create_sop(item_id, f'"{sops}"', item_name.replace('"', ''), folder_selected + '/')
        else:
            self.create_sop(item_id, f'"{sops}"', original_name.replace('"', ''), folder_selected + '/')
    
    def search_in_sop(self, item_name, folder_selected, sop):
        item_quantity = 0
        for dirpath, dirnames, filenames in os.walk(folder_selected):
            for file in filenames:
                if sop in file.split('_'):
                    try:
                        wb = xl.load_workbook(dirpath + "/" + file, data_only=True)
                        sheet = wb[wb.sheetnames[0]]

                        materials_row = None
                        tools_row = None
                        quantities_col = None
                        for row in range(1, sheet.max_row):
                            if sheet.cell(row, 1).value == "Materiais /Materials":
                                materials_row = row
                                break
                        
                        for row in range(1, sheet.max_row):
                            if sheet.cell(row, 1).value == "Ferramentas/ Tools":
                                tools_row = row
                                break

                        for col in range(1, sheet.max_column):
                            if sheet.cell(materials_row, col).value == "Quant. [units]":
                                quantities_col = col
                                break
                        for row in range(materials_row+1, tools_row):
                            if sheet.cell(row, quantities_col).value != None:
                                if sheet.cell(row, 1).value == item_name:
                                    item_quantity = sheet.cell(row, quantities_col).value
                                    break
                    except:
                        pass
            break
        return sop + ' ( ' + str(item_quantity) + ' )'

    def get_items_in_group(self, board_id, group_id):
        items = []
        item_names = []
        query = "query { boards " + f"(ids: {board_id}) " + "{ groups " + f"(ids: {group_id}) " + "{ items { id name } } } }"
        data = {'query': query}

        response = requests.post(self.base_url, headers=self.headers, json=data)

        for item in response.json()['data']['boards'][0]['groups'][0]['items']:
            items.append(item['id'])
            item_names.append(item['name'].lower())
        
        return items, item_names

    def get_current_value(self, item_id, column_id):
        query = "query { items " + f"(ids: [{item_id}]) " + "{ column_values " + f"(ids: {column_id})" + "{ value text } } }"
        data = {'query': query}

        response1 = requests.post(self.base_url, headers=self.headers, json=data)

        return response1.json()['data']['items'][0]['column_values'][0]['text']
    
    def set_status(self, column_id, item_id, board_id):
        query = 'mutation { change_simple_column_value ' + f'(item_id: {item_id}, board_id: {board_id}, column_id: {column_id}, value: "Inserido Automaticamente" )' + " { id } }"
        data = {'query': query}

        response = requests.post(self.base_url, headers=self.headers, json=data)
    
    def add_box(self, board_id, item_id, box_id, value):
        self.change_column_value(board_id, item_id, box_id, value)
    def add_time(self, board_id, item_id, time_id, value):
        self.change_column_value(board_id, item_id, time_id, value)
    def add_mo(self, board_id, item_id, mo_id, value):
        self.change_column_value(board_id, item_id, mo_id, value)
    def add_sop(self, board_id, item_id, sop_id, value):
        self.change_column_value(board_id, item_id, sop_id, value)

    def procedures(self, board_id, group_id, box_id, time_id, mo_id, sop_id, item_name, value):
        item_id = self.create_item(board_id, group_id, item_name)
        self.add_box(board_id, item_id, box_id, value[0])
        self.add_time(board_id, item_id, time_id, value[1])
        self.add_mo(board_id, item_id, mo_id, value[2])
        self.add_sop(board_id, item_id, sop_id, value[3])
    



m = MondayAPI()
#m.set_status('status_1', '3192203140', '3192202763')
#m.get_current_quantity('3192203033', 'n_meros5')
#m.search_in_sop('rolos','', 'SOP022')
# print(m.get_board_id('https://composite-solutions.monday.com/boards/3192202763'))
#item_id = m.create_item('3192202763', 'novo_grupo20271', '"new item"')
#print(item_id)
#m.change_column_value('3192202763', item_id, 'n_meros5', 2)
#m.create_sop(item_id, '"SOP021, SOP022, SOP023, SOP024, SOP025, SOP028"', 'catalisador','')
#m.get_items_in_group('3192202763', 'novo_grupo20271')
