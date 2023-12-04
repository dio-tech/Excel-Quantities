import os
import sys
from create_popup import popup_showinfo
from info import Info
import openpyxl as xl
from tkinter import filedialog
from tkinter import *
from tqdm import tqdm
from tabulate import tabulate
from tkinter import ttk

from global_functions import *
from monday_api import MondayAPI
from procedures import add_procedures

root = Tk()
folder_selected = filedialog.askdirectory(title="Escolher pasta com os procedimentos")
root.destroy()
root.quit()

materials = []

for dirpath, dirnames, filenames in os.walk(folder_selected):
    print("Getting materials...")
    for file in tqdm(filenames):
        try:
            wb = xl.load_workbook(dirpath + "/" + file)
            sheet = wb['Condições Iniciais']

            materials_row, tools_row, quantities_col = get_rows(sheet)

            for row in range(materials_row+1, tools_row):
                if sheet.cell(row, 1).value != None and sheet.cell(row, 1).value != "Materiais /Materials":
                    materials.append(sheet.cell(row, 1).value.lower())
        except:
            pass
    
    materials = list(sorted(list(set(materials))))
    materials_quantities = [0 for _ in range(len(materials))]
    sops = [[] for _ in range(len(materials))]
    print("Getting quantities...")
    for file in tqdm(filenames):
        try:
            wb = xl.load_workbook(dirpath + "/" + file, data_only=True)
            sheet = wb['Condições Iniciais']
            materials_row, tools_row, quantities_col = get_rows(sheet)
            
            for row in range(materials_row+1, tools_row-1):
                if sheet.cell(row, 1).value != None:
                    index = materials.index(sheet.cell(row, 1).value.lower())
                    materials_quantities[index] = add_quantity(sheet.cell(row, 1).value, materials_quantities, sheet.cell(row, quantities_col).value, index)
                    sops[index].append(get_sop(file))
        except:
            pass

    break

if not os.path.exists(folder_selected + '/' + 'Quantities'):
    os.mkdir(folder_selected + '/' + 'Quantities')

quantities_path = folder_selected + '/' + 'Quantities'

workbook = xl.Workbook()
sheet = workbook.active

for i in range(len(materials)):
    total_sop = ''
    for index, sop in enumerate(sops[i]):
        if index != len(sops[i]) - 1:
            total_sop += str(sop) + ', '
        else:
            total_sop += str(sop)
    sheet['A' + str(i+1)] = materials[i]
    sheet['B' + str(i+1)] = materials_quantities[i]
    sheet['C' + str(i+1)] = total_sop

sheet.column_dimensions['A'].width = 40
sheet.column_dimensions['B'].width = 20
sheet.column_dimensions['C'].width = 50

workbook.save(quantities_path + '/' + 'quantities.xlsx')

original_info = []
info = []

for i in range(len(materials)):
    total_sop = ''
    for index, sop in enumerate(sops[i]):
        if index != len(sops[i]) - 1:
            total_sop += str(sop) + ', '
        else:
            total_sop += str(sop)
    original_info.append([materials[i], materials_quantities[i], total_sop])
    info.append([materials[i], materials_quantities[i], total_sop])

print(tabulate(original_info, headers=["Material", "Quantity", "SOP"]))

if len(original_info) > 0:
    in1 = Info()
    choice = in1.get_choice()

    monday = MondayAPI()

    board_id = ""
    groups_name = []
    columns_name = []
    group_chosen = [0]
    column_chosen = [0]
    unit_chosen = [0]
    status = [0]

    def add():
        group_chosen[0] = groups[groupchosen.current()]
        column_chosen[0] = columns[columnchosen.current()]
        unit_chosen[0] = columns[unitchosen.current()]
        status[0] = columns[statuschosen.current()]
        root.destroy()
        root.quit()

    if choice == 1:
        pass
    
    else:
        print("Abrir ficheiro excel em: " + quantities_path + '/' + 'quantities.xlsx')
        print("Editar, guardar e fechar antes de voltar a correr o programa")
        input("Pressionar Enter para continuar (depois de editar o ficheiro)")

        file = quantities_path + '/' + 'quantities.xlsx'

        wb = xl.load_workbook(file)
        sheet = wb[wb.sheetnames[0]]

        info = []

        for row in range(1, sheet.max_row+1):
            info.append([sheet.cell(row, 1).value, sheet.cell(row, 2).value, sheet.cell(row, 3).value])

    
    b_id = input('Inserir link do quadro: ')
    board_id = monday.get_board_id(b_id)

    add_procedures(folder_selected, quantities_col, board_id)


    groups = monday.get_groups(board_id)
    for group in groups:
        groups_name.append(group[1])
        
    columns = monday.board_columns(board_id)
    for column in columns:
        columns_name.append(column[1])

    root = Tk('Insert to Monday')
    root.title('Insert to Monday (Materiais)')

    root.state('zoomed')

    b_id = Label(root, text="Escolher grupo para adicionar informação: ")
    qt_l = Label(root, text="Escolher coluna para adicionar quantidades: ")
    un_l = Label(root, text="Escolher coluna para adicionar unidades: ")
    st_l = Label(root, text="Escolher coluna para definir status do material: ")
    btn2 = Button(root, text='ADICIONAR AO MONDAY', command=add)

    # Combobox creation
    n = StringVar()
    n1 = StringVar()
    n2 = StringVar()
    n3 = StringVar()
    groupchosen = ttk.Combobox(root, width = 27, textvariable = n)
    columnchosen = ttk.Combobox(root, width = 27, textvariable = n1)
    unitchosen = ttk.Combobox(root, width = 27, textvariable = n2)
    statuschosen = ttk.Combobox(root, width = 27, textvariable = n3)

    groupchosen['values'] = groups_name
    columnchosen['values'] = columns_name
    unitchosen['values'] = columns_name
    statuschosen['values'] = columns_name

    b_id.pack()
    groupchosen.pack(pady=10)
    qt_l.pack(pady=10)
    columnchosen.pack(pady=10)
    un_l.pack(pady=10)
    unitchosen.pack(pady=10)
    st_l.pack(pady=10)
    statuschosen.pack(pady=10)
    btn2.pack(pady=10)

    root.mainloop()
    root.quit()

    group_chosen = group_chosen[0]
    column_chosen = column_chosen[0]
    unit_chosen = unit_chosen[0]
    status = status[0]

    print("Adding...")
    for i in tqdm(info):
        items_id, item_names = monday.get_items_in_group(board_id, group_chosen[0])
        if i[0].lower() not in item_names:
            monday.add_to_monday(board_id, group_chosen[0], f'"{i[0]}"', column_chosen[0], unit_chosen[0], status[0], i[1], i[2], folder_selected)
        else:
            popup_response = popup_showinfo(i[0])
            if popup_response == 2:
                item_id = items_id[item_names.index(i[0])]
                monday.add_to_monday(board_id, group_chosen[0], f'"{i[0]}"', column_chosen[0], unit_chosen[0], status[0], i[1], i[2], folder_selected, create=False, item_id=item_id)
            else:
                item_id = items_id[item_names.index(i[0])]
                current_quantity = monday.get_current_value(item_id, column_chosen[0])
                unit = monday.get_current_value(items_id[item_names.index(i[0])], unit_chosen[0])
                quantity = i[1].split(' ')
                quantity[0] = str(float(quantity[0]) + float(current_quantity))
                quantity[1] = str(unit)
                q = ' '.join(quantity)
                monday.add_to_monday(board_id, group_chosen[0], f'"{i[0]}"', column_chosen[0], unit_chosen[0], status[0], q, i[2], folder_selected, create=False, item_id=item_id, original_name=original_info[info.index(i)][0])

