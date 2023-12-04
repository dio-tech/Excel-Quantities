import os
from tqdm import tqdm
import openpyxl as xl
from create_popup import popup_for_missing_requirements
from monday_api import *
from tkinter import *
from tkinter import ttk

def add_procedures(folder_selected, quantities_column, board_id):
    box = []
    time = []
    mo = []
    name = []
    sop = []
    for dirpath, dirnames, filenames in os.walk(folder_selected):
        print("Getting procedures...")
        for index, file in tqdm(enumerate(filenames)):
            try:
                for i in file.split('_'):
                    if i.startswith('SOP'):
                        sop.append(i.replace('SOP', ''))
                wb = xl.load_workbook(dirpath + "/" + file)
                sheet = wb['Condições Iniciais']
                name.append(sheet.cell(2, 3).value)

                for i in range(1, sheet.max_row):
                    if sheet.cell(i, 1).value == "Box":
                        if sheet.cell(i, quantities_column) != None:
                            box.append(sheet.cell(i, quantities_column).value)
                    elif sheet.cell(i, 1).value == "Tempo Previsto/ Estimated time [h]":
                        time.append(sheet.cell(i, quantities_column).value)
                    elif sheet.cell(i, 1).value == "Mão de obra [M.O]":
                        mo.append(sheet.cell(i, quantities_column).value)
                
                if len(box) - 1 < index:
                    box.append(popup_for_missing_requirements('Box', dirpath+file))
                if len(mo) - 1 < index:
                    mo.append(popup_for_missing_requirements('Mão de obra [M.O]', dirpath+file))
                if len(time) - 1 < index:
                    time.append("Tempo Previsto/ Estimated time [h]", dirpath+file)

            except:
                pass
        break
    
    monday = MondayAPI()
    columns = monday.board_columns(board_id)
    columns_name = []
    for column in columns:
        columns_name.append(column[1])
    
    groups = monday.get_groups(board_id)
    groups_name = []
    for group in groups:
        groups_name.append(group[1])
    
    box_id = [0]
    time_id = [0]
    mo_id = [0]
    group_id = [0]
    sop_id = [0]
    
    def add():
        group_id[0] = groups[group_chosen.current()]
        box_id[0] = columns[box_chosen.current()]
        time_id[0] = columns[time_chosen.current()]
        mo_id[0] = columns[m_o.current()]
        sop_id[0] = columns[sop_chosen.current()]
        root.destroy()
        root.quit()
    
    root = Tk('Insert to Monday')
    root.title('Insert to Monday (Procedimentos)')

    root.state('zoomed')

    g_id = Label(root, text="Escolher grupo para adicionar informação: ")
    b_id = Label(root, text="Escolher coluna para adicionar Box: ")
    qt_l = Label(root, text="Escolher coluna para adicionar Tempo Previsto/ Estimated time [h]: ")
    un_l = Label(root, text="Escolher coluna para adicionar Mão de obra [M.O]: ")
    s_l = Label(root, text="Escolher coluna para adicionar SOP: ")
    btn2 = Button(root, text='ADICIONAR AO MONDAY', command=add)

    # Combobox creation
    n = StringVar()
    n1 = StringVar()
    n2 = StringVar()
    n3 = StringVar()
    n4 = StringVar()

    box_chosen = ttk.Combobox(root, width = 27, textvariable = n)
    time_chosen = ttk.Combobox(root, width = 27, textvariable = n1)
    m_o = ttk.Combobox(root, width = 27, textvariable = n2)
    group_chosen = ttk.Combobox(root, width = 27, textvariable = n3)
    sop_chosen = ttk.Combobox(root, width = 27, textvariable = n4)

    box_chosen['values'] = columns_name
    time_chosen['values'] = columns_name
    m_o['values'] = columns_name
    sop_chosen['values'] = columns_name
    group_chosen['values'] = groups_name

    g_id.pack(pady=10)
    group_chosen.pack(pady=10)
    b_id.pack(pady=10)
    box_chosen.pack(pady=10)
    qt_l.pack(pady=10)
    time_chosen.pack(pady=10)
    un_l.pack(pady=10)
    m_o.pack(pady=10)
    s_l.pack(pady=10)
    sop_chosen.pack(pady=10)
    btn2.pack(pady=10)

    root.mainloop()
    root.quit()

    print('Adding procedures...')
    for k in tqdm(range(len(name))):
        monday.procedures(board_id, group_id[0][0], box_id[0][0], time_id[0][0], mo_id[0][0], sop_id[0][0], f'"{name[k]}"', [box[k], time[k], mo[k], sop[k]])

#add_procedures('C:/Users/diogo/OneDrive/Ambiente de Trabalho/procedimentos/', 6, '3192202763')